import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import datetime
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Desteklenen varsayılan PDF fontları (Türkçe karakter sorunu için)
# Eger sistemde Arial varsa Arial kullanabiliriz, yoksa default Helvetica mecburi kullanilir ama turkce karakter desteklemez
# Gecici olarak reportlab in kendi fontunu veya kayitsiz UTF-8 destegini koymaliyiz.

def generate_batch_excel(invoices: list, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Toplu Yevmiye Fişi"

    # Baslik Stilleri
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    headers = ["Tarih", "Evrak No", "Firma Adı", "Hesap Kodu", "Hesap Adı", "Borç (TL)", "Alacak (TL)", "Açıklama"]
    
    # Basliklari yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = align_center

    row_num = 2
    toplam_borc = 0.0
    toplam_alacak = 0.0

    # Her fatura icin kayitlari olustur
    for inv in invoices:
        date_str = inv.invoice_date or ""
        inv_no = inv.invoice_number or ""
        company = inv.company_name or ""
        
        # 1. GİDER SATIRI (Borç)
        base = inv.base_amount or 0.0
        ws.append([date_str, inv_no, company, inv.expense_account, "Giderler Hesabı", base, 0.0, f"Matrah - {inv_no}"])
        toplam_borc += base

        # 2. KDV SATIRI (Borç)
        vat = inv.vat_amount or 0.0
        if vat > 0:
            ws.append([date_str, inv_no, company, inv.vat_account, "İndirilecek KDV", vat, 0.0, f"KDV (%{inv.vat_rate}) - {inv_no}"])
            toplam_borc += vat

        # 3. YÜKÜMLÜLÜK SATIRI (Alacak)
        total = inv.total_amount or 0.0
        ws.append([date_str, inv_no, company, inv.vendor_account, "Satıcılar", 0.0, total, f"Genel Toplam - {company}"])
        toplam_alacak += total

        # Satirlara kenarlik ekle
        for r in range(row_num, row_num + (3 if vat > 0 else 2)):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = border

        row_num += (3 if vat > 0 else 2)

    # Toplam Satırı
    ws.append(["", "", "", "", "GENEL TOPLAM:", toplam_borc, toplam_alacak, ""])
    for c in range(5, 8):
        cell = ws.cell(row=row_num, column=c)
        cell.font = Font(bold=True)
        cell.border = border
        if c in [6, 7]:
            cell.alignment = align_right

    # Sutun genisliklerini ayarla
    col_widths = {"A": 12, "B": 20, "C": 35, "D": 12, "E": 20, "F": 15, "G": 15, "H": 40}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Formatlama
    for row in range(2, row_num + 1):
        ws.cell(row, 6).number_format = '#,##0.00'
        ws.cell(row, 7).number_format = '#,##0.00'

    wb.save(out_path)
    return out_path


def generate_batch_pdf(invoices: list, out_path: str):
    doc = SimpleDocTemplate(out_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Heading1'],
        alignment=1, # Center
        spaceAfter=20
    )
    
    # Baslik
    date_now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    elements.append(Paragraph(f"TOPLU YEVMIYE FISI (BATCH JOURNAL VOUCHER)", title_style))
    elements.append(Paragraph(f"Olusturulma Tarihi: {date_now}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Tablo verisi
    data = [["Tarih", "Evrak No", "Firma Adi", "Hesap", "Borc (TL)", "Alacak (TL)"]]
    
    toplam_borc = 0.0
    toplam_alacak = 0.0

    for inv in invoices:
        date_str = inv.invoice_date or ""
        # Cok uzun isimleri kes
        company = (inv.company_name or "")[:20]
        inv_no = (inv.invoice_number or "")
        
        # Gider
        base = inv.base_amount or 0.0
        data.append([date_str, inv_no, company, str(inv.expense_account), f"{base:,.2f}", ""])
        toplam_borc += base
        
        # KDV
        vat = inv.vat_amount or 0.0
        if vat > 0:
            data.append(["", "", "", str(inv.vat_account), f"{vat:,.2f}", ""])
            toplam_borc += vat

        # Cari
        total = inv.total_amount or 0.0
        data.append(["", "", "", str(inv.vendor_account), "", f"{total:,.2f}"])
        toplam_alacak += total

    # Genel Toplam
    data.append(["", "", "", "GENEL TOPLAM", f"{toplam_borc:,.2f}", f"{toplam_alacak:,.2f}"])

    # Tablo Stili
    t = Table(data, colWidths=[60, 100, 150, 50, 80, 80])
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (4, 1), (5, -1), 'RIGHT'), # Tutarlari saga yasla
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), # Toplam satiri bold
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    return out_path
